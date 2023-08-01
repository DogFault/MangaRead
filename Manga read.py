import openpyxl
from googlesearch import search
from datetime import datetime
import pandas as pd
import requests
from bs4 import BeautifulSoup

def search_manga_online(manga_name):
    try:
        search_query = f"{manga_name} manga update"
        search_results = search(search_query)
        most_viewed_result = None
        max_views = -1

        for result in search_results:
            print(result)
            views = get_view_count(result)  # Get the view count (replace this with the actual method to get view count)
            print(f"Views: {views}")
            if views is None:  # If view count is None, set it to 0
                views = 0

            if views > max_views:
                max_views = views
                most_viewed_result = result

        if most_viewed_result:
            print(f"Most viewed search result for '{manga_name}' update:")
            print(most_viewed_result)
            data = {
                "Manga Name": [manga_name],
                "Date": [datetime.now()],
                "Website": [most_viewed_result],
                "Views": [max_views],
                "Status": ["Good" if max_views > 0 else "Bad"]
            }
            return pd.DataFrame(data)
        else:
            print(f"No valid update search result found for '{manga_name}'.")
            data = {
                "Manga Name": [manga_name],
                "Date": [datetime.now()],
                "Website": [None],
                "Views": [None],
                "Status": ["Bad"]
            }
            return pd.DataFrame(data)  # Return a DataFrame with default values
    except Exception as e:
        print(f"An error occurred while searching for '{manga_name}': {e}")
        return None  # Return None when an error occurs

# Helper method to get the view count from a search result URL (replace this with actual implementation)
def get_view_count(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an error for unsuccessful HTTP status codes
        if response.status_code == 200:
            # Parse the response content using BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            # Replace 'view_count_element_class' with the actual class name of the HTML element containing the view count
            view_count_element = soup.find(class_='view_count_element_class')
            if view_count_element:
                view_count = int(view_count_element.get_text())
            else:
                view_count = 0  # Set a default value when the view count element is not found
            return view_count
        else:
            print(f"Failed to fetch view count from {url}. Status code: {response.status_code}")
            return None  # Return None when the request is unsuccessful or there is no view count element
    except requests.exceptions.HTTPError as http_err:
        if response.status_code in (403, 406):
            print(f"HTTP error occurred while fetching view count from {url}: {http_err}")
            return None  # Return None when encountering specific HTTP errors (403 or 406)
        else:
            raise  # Re-raise other HTTP errors
    except Exception as e:
        print(f"Error occurred while fetching view count from {url}: {e}")
        return None  # Return None when an error occurs during the fetching process


if __name__ == "__main__":
    # Load the existing workbook and get the desired sheet by name
    excel_file_path = "manga_list.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet_name = "Sheet1"  # Replace "Sheet1" with the actual name of your sheet
    sheet = workbook[sheet_name]

    # Get manga names from the Excel sheet
    manga_names = [cell.value for cell in sheet['A'][1:]]  # Assuming manga names are in the first column (column A)

    # Perform manga search and get the sorted results
    # List to store the results of each manga search
    search_results_list = []

    # Iterate through the manga names and perform the search
    for manga_name in manga_names:
        print(f"Searching for updates on '{manga_name}'...")
        search_results = search_manga_online(manga_name)
        search_results_list.append(search_results)

    # Concatenate the search results and sort by date
    sorted_results = pd.concat(search_results_list, ignore_index=True).sort_values(by="Date", ascending=False)

    # Update the Excel file with the sorted search results and status column
    for index, row in sorted_results.iterrows():
        manga_name = row["Manga Name"]
        manga_row = index + 2  # Adjust the index by 2 to account for Excel rows starting at 2
        sheet.cell(row=manga_row, column=2, value=manga_name)
        sheet.cell(row=manga_row, column=3, value=row["Date"].strftime("%Y-%m-%d %H:%M:%S"))
        sheet.cell(row=manga_row, column=4, value=row["Website"])
        if row["Views"] is not None and row["Views"] > 0:
            sheet.cell(row=manga_row, column=5, value=row["Views"])
        else:
            sheet.cell(row=manga_row, column=5, value="")  # If view count is None or 0, leave the cell empty
        sheet.cell(row=manga_row, column=6, value=row["Status"])  # Update the Status column

    # Save the updated workbook
    workbook.save(excel_file_path)
